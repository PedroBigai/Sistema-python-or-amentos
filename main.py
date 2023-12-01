from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from openpyxl import load_workbook
import re


class AplicativoPersiana(App):
    def build(self):
        self.layout = BoxLayout(orientation='vertical', padding=20, spacing=15)
        
        self.largura_input = TextInput(hint_text="Largura", font_size=20, foreground_color=(1, 1, 1), background_color=(0, 0, 0), padding=10)
        self.layout.add_widget(self.largura_input)
        
        self.altura_input = TextInput(hint_text="Altura", font_size=20, foreground_color=(1, 1, 1), background_color=(0, 0, 0))
        self.layout.add_widget(self.altura_input)

        self.codigo_input = TextInput(hint_text="Qual o código do tecido:", font_size=20, foreground_color=(1, 1, 1), background_color=(0, 0, 0))
        self.layout.add_widget(self.codigo_input)

        self.modelo_persiana_input = Button(text="Modelo", on_press=self.abrir_popup_modelos)
        self.layout.add_widget(self.modelo_persiana_input)

        self.acess_button = Button(text="Acessórios", on_press=self.abrir_janela_acessorios)
        self.layout.add_widget(self.acess_button)

        
        self.resultado_label = Label(text="", font_size=20, size_hint_y=5, height=100)
        self.layout.add_widget(self.resultado_label)
        
        return self.layout
        
    def abrir_popup_modelos(self, instance):
        from kivy.uix.popup import Popup
        content = BoxLayout(orientation='vertical', spacing=10, padding=20)
        botão_rolo = Button(text="Rolô", on_press=self.modelo_rolo)
        content.add_widget(botão_rolo)
        botão_romana = Button(text="Romana", on_press=self.modelo_romana)
        content.add_widget(botão_romana)
        botão_ph = Button(text="PH", on_press=self.modelo_ph)
        content.add_widget(botão_ph)
        botão_dv = Button(text="DV Screen", on_press=self.modelo_dv)
        content.add_widget(botão_dv)
        botão_pv = Button(text="PV", on_press=self.modelo_pv)
        content.add_widget(botão_pv)
        botão_pv_bk = Button(text="PV BK", on_press=self.modelo_pv_bk)
        content.add_widget(botão_pv_bk)
        self.popup = Popup(title="Escolher o modelo", content=content, size_hint=(None, None), size=(600, 600))
        self.popup.open()
    

    def modelo_rolo(self, instance) :
        modelo_persiana = "Rolô"
        self.modelo_persiana_input.text = modelo_persiana
        try:
            largura = float(self.largura_input.text)
            altura = float(self.altura_input.text)
        except ValueError:
            self.resultado_label.text = "Digite números válidos."
            self.popup.dismiss()
            return
        resultado = altura * largura
        formatacao = "{:.2f}".format(resultado)

        modelo_persiana = self.modelo_persiana_input.text
        perguntar_tecido = self.codigo_input.text
        
        if modelo_persiana == "Rolô":
            planilha = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Rolo.xlsx")  # Substitua pelo caminho correto
            planilha_rolo = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Tabela Jô Decorações - ROLO + DV + SCREEN + BK (1).xlsx")  # Substitua pelo caminho correto
            ws_rolo = planilha["Table 1"]
            ws_rolo2 = planilha_rolo["Table 2"]
            for row in ws_rolo.iter_rows(min_row=3, values_only=True, min_col=0, max_col=10, max_row=39):  # Substitua o número de linhas máximo conforme necessário
                tecido = row[0]
                codigo = row[1]
                largura = row[2]
                preco_vista = row[6]
                preco_prazo = row[8]
                if perguntar_tecido in str(codigo) or perguntar_tecido == str(tecido):
                    self.resultado_label.text = f"Tecido encontrado na tabela:\n"
                    self.resultado_label.text += f"Tecido: {tecido}\n"
                    self.resultado_label.text += f"Código: {codigo}\n"
                    self.resultado_label.text += f"Larguta: {largura}\n"
                    self.resultado_label.text += f"Preço à vista: R${preco_vista:.2f}\n"
                    self.resultado_label.text += f"Preço a prazo: R${preco_prazo:.2f}\n"
                    self.resultado_label.text += f"Total de tecido: {formatacao}m\n"

                    preco_total_vista = (preco_vista * resultado) * 1.5
                    preco_total_prazo = (preco_prazo * resultado) * 1.5

                    self.resultado_label.text += f"Preço total à vista: R${preco_total_vista:.2f}\n"
                    self.resultado_label.text += f"Preço total a prazo: R${preco_total_prazo:.2f}\n"

                    self.total_vista_acessorios =  preco_total_vista
                    self.total_prazo_acessorios =  preco_total_prazo

            for row in ws_rolo2.iter_rows(min_row=3, values_only=True, min_col=0, max_col=10, max_row=39):  # Substitua o número de linhas máximo conforme necessário
                tecido = row[0]
                codigo = row[1]
                largura = row[2]
                preco_vista = row[4]
                preco_prazo = row[6]
                if perguntar_tecido in str(codigo) or perguntar_tecido == str(tecido):
                    self.resultado_label.text = f"Tecido encontrado na tabela:\n"
                    self.resultado_label.text += f"Tecido: {tecido}\n"
                    self.resultado_label.text += f"Código: {codigo}\n"
                    self.resultado_label.text += f"Larguta: {largura}\n"
                    self.resultado_label.text += f"Preço à vista: R${preco_vista:.2f}\n"
                    self.resultado_label.text += f"Preço a prazo: R${preco_prazo:.2f}\n"
                    self.resultado_label.text += f"Total de tecido: {formatacao}m\n"

                    preco_total_vista = (preco_vista * resultado) * 1.5
                    preco_total_prazo = (preco_prazo * resultado) * 1.5

                    self.resultado_label.text += f"Preço total à vista: R${preco_total_vista:.2f}\n"
                    self.resultado_label.text += f"Preço total a prazo: R${preco_total_prazo:.2f}\n"

                    self.total_vista_acessorios =  preco_total_vista
                    self.total_prazo_acessorios =  preco_total_prazo
                    break
        self.popup.dismiss()

    def modelo_romana(self, instance) :
        self.modelo_persiana_input.text = "Romana"
        try:
            largura = float(self.largura_input.text)
            altura = float(self.altura_input.text)
        except ValueError:
            self.resultado_label.text = "Digite números válidos."
            self.popup.dismiss()
            return
        resultado = altura * largura
        formatacao = "{:.2f}".format(resultado)

        modelo_persiana = self.modelo_persiana_input.text
        perguntar_tecido = self.codigo_input.text
        if modelo_persiana == "Romana":
            planilha6 = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Romana.xlsx")  # Substitua pelo caminho correto
            ws_romana = planilha6["Table 1"]
            for row in ws_romana.iter_rows(min_row=4, values_only=True, min_col=0, max_col=10, max_row=49):  # Substitua o número de linhas máximo conforme necessário
                tecido = row[0]
                codigo = row[1]
                largura = row[4]
                preco_vista = row[6]
                preco_prazo = row[8]
                if perguntar_tecido in str(codigo) or perguntar_tecido == str(tecido):
                    self.resultado_label.text = f"Tecido encontrado na tabela:\n"
                    self.resultado_label.text += f"Tecido: {tecido}\n"
                    self.resultado_label.text += f"Código: {codigo}\n"
                    self.resultado_label.text += f"Larguta: {largura}\n"
                    self.resultado_label.text += f"Preço à vista: R${preco_vista:.2f}\n"
                    self.resultado_label.text += f"Preço a prazo: R${preco_prazo:.2f}\n"
                    self.resultado_label.text += f"Total de tecido: {formatacao}m\n"

                    preco_total_vista = (preco_vista * resultado) * 1.5
                    preco_total_prazo = (preco_prazo * resultado) * 1.5

                    self.resultado_label.text += f"Preço total à vista: R${preco_total_vista:.2f}\n"
                    self.resultado_label.text += f"Preço total a prazo: R${preco_total_prazo:.2f}\n"

                    self.total_vista_acessorios =  preco_total_vista
                    self.total_prazo_acessorios =  preco_total_prazo
                    break
                else:
                    self.resultado_label.text = "Modelo/Código de cortina não encontrado."
        self.popup.dismiss()

    def modelo_ph(self, instance) :
        self.modelo_persiana_input.text = "PH"
        try:
            largura = float(self.largura_input.text)
            altura = float(self.altura_input.text)
        except ValueError:
            self.resultado_label.text = "Digite números válidos."
            self.popup.dismiss()
            return
        resultado = altura * largura
        formatacao = "{:.2f}".format(resultado)

        modelo_persiana = self.modelo_persiana_input.text
        perguntar_tecido = self.codigo_input.text
        if modelo_persiana == "PH":
            planilha2 = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/PH.xlsx")  # Substitua pelo caminho correto
            ws_ph = planilha2["Table 1"]
            for row in ws_ph.iter_rows(min_row=2, values_only=True, min_col=0, max_col=7, max_row=38):  # Substitua o número de linhas máximo conforme necessário
                tecido = row[0]
                codigo = row[1]
                preco_vista = row[5]
                preco_prazo = row[6]
                if perguntar_tecido in str(codigo) or perguntar_tecido == str(tecido):
                    self.resultado_label.text = f"Tecido encontrado na tabela:\n"
                    self.resultado_label.text += f"Tecido: {tecido}\n"
                    self.resultado_label.text += f"Código: {codigo}\n"

                    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
                    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())


                    self.resultado_label.text += f"Preço à vista: R${preco_vista:.2f}\n"
                    self.resultado_label.text += f"Preço a prazo: R${preco_prazo:.2f}\n"
                    self.resultado_label.text += f"Total de tecido: {formatacao}m\n"
                
                    preco_total_vista = (preco_vista * resultado) * 1.5
                    preco_total_prazo = (preco_prazo * resultado) * 1.5

                    self.resultado_label.text += f"Preço total à vista: R${preco_total_vista:.2f}\n"
                    self.resultado_label.text += f"Preço total a prazo: R${preco_total_prazo:.2f}\n"

                    self.total_vista_acessorios =  preco_total_vista
                    self.total_prazo_acessorios =  preco_total_prazo
                    break 
                else:
                    self.resultado_label.text = "Modelo/Código de cortina não encontrado."
        self.popup.dismiss()

    def modelo_dv(self, instance) :
        self.modelo_persiana_input.text = "DV"
        try:
            largura = float(self.largura_input.text)
            altura = float(self.altura_input.text)
        except ValueError:
            self.resultado_label.text = "Digite números válidos."
            self.popup.dismiss()
            return
        resultado = altura * largura
        formatacao = "{:.2f}".format(resultado)

        modelo_persiana = self.modelo_persiana_input.text
        perguntar_tecido = self.codigo_input.text
        if modelo_persiana == "DV":
            planilha5 = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Rolo DV Screen.xlsx")  # Substitua pelo caminho correto
            ws_dv = planilha5["Table 2"]
            for row in ws_dv.iter_rows(min_row=1, values_only=True, min_col=0, max_col=10, max_row=39):  # Substitua o número de linhas máximo conforme necessário
                tecido = row[0]
                codigo = row[1]
                largura = row[2]
                preco_vista = row[4]
                preco_prazo = row[6]
                if perguntar_tecido in str(codigo) or perguntar_tecido == str(tecido):
                    self.resultado_label.text = f"Tecido encontrado na tabela:\n"
                    self.resultado_label.text += f"Tecido: {tecido}\n"
                    self.resultado_label.text += f"Código: {codigo}\n"
                    self.resultado_label.text += f"Largura: {largura}\n"
                    self.resultado_label.text += f"Preço à vista: R${preco_vista:.2f}\n"
                    self.resultado_label.text += f"Preço a prazo: R${preco_prazo:.2f}\n"
                    self.resultado_label.text += f"Total de tecido: {formatacao}m\n"

                    preco_total_vista = (preco_vista * resultado) * 1.5
                    preco_total_prazo = (preco_prazo * resultado) * 1.5

                    self.resultado_label.text += f"Preço total à vista: R${preco_total_vista:.2f}\n"
                    self.resultado_label.text += f"Preço total a prazo: R${preco_total_prazo:.2f}\n"

                    self.total_vista_acessorios =  preco_total_vista
                    self.total_prazo_acessorios =  preco_total_prazo
                    break
                else:
                    self.resultado_label.text = "Modelo/Código de cortina não encontrado."
        self.popup.dismiss()

    def modelo_pv(self, instance) :
        self.modelo_persiana_input.text = "PV"
        try:
            largura = float(self.largura_input.text)
            altura = float(self.altura_input.text)
        except ValueError:
            self.resultado_label.text = "Digite números válidos."
            self.popup.dismiss()
            return
        resultado = altura * largura
        formatacao = "{:.2f}".format(resultado)

        modelo_persiana = self.modelo_persiana_input.text
        perguntar_tecido = self.codigo_input.text
        if modelo_persiana == "PV":
            planilha3 = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Pv.xlsx")  # Substitua pelo caminho correto
            ws_pv = planilha3["Table 1"]
            for row in ws_pv.iter_rows(min_row=2, values_only=True, min_col=0, max_col=7, max_row=38):  # Substitua o número de linhas máximo conforme necessário
                tecido = row[0]
                codigo = row[1]
                preco_vista = row[5]
                preco_prazo = row[6]
                if perguntar_tecido in str(codigo) or perguntar_tecido == str(tecido):
                    self.resultado_label.text = f"Tecido encontrado na tabela:\n"
                    self.resultado_label.text += f"Tecido: {tecido}\n"
                    self.resultado_label.text += f"Código: {codigo}\n"

                    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
                    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())


                    self.resultado_label.text += f"Preço à vista: R${preco_vista:.2f}\n"
                    self.resultado_label.text += f"Preço a prazo: R${preco_prazo:.2f}\n"
                    self.resultado_label.text += f"Total de tecido: {formatacao}m\n"
                
                    preco_total_vista = (preco_vista * resultado) * 1.5
                    preco_total_prazo = (preco_prazo * resultado) * 1.5

                    self.resultado_label.text += f"Preço total à vista: R${preco_total_vista:.2f}\n"
                    self.resultado_label.text += f"Preço total a prazo: R${preco_total_prazo:.2f}\n"

                    self.total_vista_acessorios =  preco_total_vista
                    self.total_prazo_acessorios =  preco_total_prazo
                    break 
                else:
                    self.resultado_label.text = "Modelo/Código de cortina não encontrado."
        self.popup.dismiss()

    def modelo_pv_bk(self, instance) :
        self.modelo_persiana_input.text = "PV BK"
        try:
            largura = float(self.largura_input.text)
            altura = float(self.altura_input.text)
        except ValueError:
            self.resultado_label.text = "Digite números válidos."
            self.popup.dismiss()
            return
        resultado = altura * largura
        formatacao = "{:.2f}".format(resultado)

        modelo_persiana = self.modelo_persiana_input.text
        perguntar_tecido = self.codigo_input.text
        if modelo_persiana == "PV BK":
            planilha4 = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/PV BK.xlsx")  # Substitua pelo caminho correto
            ws_pv_bk = planilha4["Table 2"]           
            for row in ws_pv_bk.iter_rows(min_row=2, values_only=True, min_col=0, max_col=7, max_row=17):  # Substitua o número de linhas máximo conforme necessário
                tecido = row[0]
                codigo = row[1]
                preco_vista = row[3]
                preco_prazo = row[5]
                if perguntar_tecido in str(codigo) or perguntar_tecido == str(tecido):
                    self.resultado_label.text = f"Tecido encontrado na tabela:\n"
                    self.resultado_label.text += f"Tecido: {tecido}\n"
                    self.resultado_label.text += f"Código: {codigo}\n"

                    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
                    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())
                    
                    self.resultado_label.text += f"Preço à vista: R${preco_vista:.2f}\n"
                    self.resultado_label.text += f"Preço a prazo: R${preco_prazo:.2f}\n"
                    self.resultado_label.text += f"Total de tecido: {formatacao}m\n"
                    preco_total_vista = preco_vista * resultado
                    preco_total_prazo = preco_prazo * resultado

                    self.resultado_label.text += f"Preço total à vista: R${preco_total_vista:.2f}\n"
                    self.resultado_label.text += f"Preço total a prazo: R${preco_total_prazo:.2f}\n"

                    self.total_vista_acessorios =  preco_total_vista
                    self.total_prazo_acessorios =  preco_total_prazo
                    break
                else:
                    self.resultado_label.text = "Modelo/Código de cortina não encontrado."
        self.popup.dismiss()

    def abrir_janela_acessorios(self, instance):
        from kivy.uix.popup import Popup
        content = BoxLayout(orientation='vertical', spacing=10, padding=20)
        self.acessorios_input = TextInput(hint_text="Acessórios", font_size=20, foreground_color=(1, 1, 1), background_color=(0, 0, 0))
        content.add_widget(self.acessorios_input)
        self.resultado_label_popup = Label(text="", font_size=16, halign="left", valign="top")
        content.add_widget(self.resultado_label_popup)
        save_button = Button(text="Buscar Acessórios", on_press=self.mostrar_acessorios)
        content.add_widget(save_button)
        save_button2 = Button(text="Adicionar", on_press=self.adicionar_acessorios)
        content.add_widget(save_button2)
        save_button3 = Button(text="Sair", on_press=self.sair)
        content.add_widget(save_button3)
        try:    
            self.resultado_label_popup.text += f"Total Preço à Vista Acessórios: R${self.total_vista_acessorios:.2f}\n"
            self.resultado_label_popup.text += f"Total Preço a Prazo Acessórios: R${self.total_prazo_acessorios:.2f}\n"

            self.popup = Popup(title="Inserir Acessórios", content=content, size_hint=(None, None), size=(600, 600))
            self.popup.open()
        except AttributeError:
            return
    def mostrar_acessorios(self, instance):
        planilha_acess = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Acess Rolo.xlsx")  # Substitua pelo caminho correto
        ws_rolo_acess = planilha_acess["Table 2"]
        planilha_acess_vertical = load_workbook("C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/PV BK.xlsx")  # Substitua pelo caminho correto
        ws_rolo_acess_vertical = planilha_acess_vertical["Table 2"]
        acessorios_digitados = self.acessorios_input.text
        encontrados = False
        resultados = []
        if acessorios_digitados:
                for row in ws_rolo_acess.iter_rows(min_row=1, values_only=True, min_col=0, max_col=6, max_row=28):  # Substitua o número de linhas máximo conforme necessário
                    acess = row[0]
                    preco_vista = row[2]
                    preco_prazo = row[4]
                    if acessorios_digitados in str(acess):
                        encontrados = True
                        resultado = f"Acessório: {acess}\n"
                        if preco_vista is not None:
                            resultado += f"A vista: R${preco_vista:.2f}\n"
                        if preco_prazo is not None:
                            resultado += f"A prazo: R${preco_prazo:.2f}\n"
                            
                            resultados.append(resultado)

                        self.preco_vista_acessorio = preco_vista if preco_vista is not None else 0.0
                        self.preco_prazo_acessorio = preco_prazo if preco_prazo is not None else 0.0
        else:
            return
               
        if encontrados:
            self.resultado_label_popup.text = "\n".join(resultados)
            
        

        for row in ws_rolo_acess_vertical.iter_rows(min_row=18, values_only=True, min_col=0, max_col=6, max_row=39):  # Substitua o número de linhas máximo conforme necessário
            acess = row[0]
            preco_vista = row[3]
            preco_prazo = row[5]
            if acessorios_digitados in str(acess):
                encontrados = True
                resultado = f"Acessório: {acess}\n"
                if preco_vista is not None:
                    resultado += f"A vista: R${preco_vista:.2f}\n"
                if preco_prazo is not None:
                    resultado += f"A prazo: R${preco_prazo:.2f}\n"
                    
                    resultados.append(resultado)

                self.preco_vista_acessorio = preco_vista if preco_vista is not None else 0.0
                self.preco_prazo_acessorio = preco_prazo if preco_prazo is not None else 0.0
                
        if encontrados:
            self.resultado_label_popup.text = "\n".join(resultados)
            
        else:
            self.resultado_label_popup.text = "Nenhum acessório encontrado."

    def adicionar_acessorios(self, instance):
    
        valor_a_vista = self.preco_vista_acessorio
        valor_a_prazo = self.preco_prazo_acessorio
        
        
        I = 0
        I2 = 0
        
        total_vista_acessorio = (valor_a_vista + I) * 1.5
        total_prazo_acessorio = (valor_a_prazo + I2) * 1.5
        
        self.total_vista_acessorios += total_vista_acessorio
        self.total_prazo_acessorios += total_prazo_acessorio
        
        self.resultado_label_popup.text = f"Total Preço à Vista C/ Acessórios: R${self.total_vista_acessorios:.2f}\n"
        self.resultado_label_popup.text += f"Total Preço a Prazo C/ Acessórios: R${self.total_prazo_acessorios:.2f}\n"

         
    def sair(self, instance):
        self.total_vista_acessorios 
        self.total_prazo_acessorios 
        self.popup.dismiss()


            
    
if __name__ == '__main__':
    AplicativoPersiana().run()
   
